from pulp import *
import openpyxl
from sympy import *
from collections import namedtuple
from math import ceil, floor


def get_values(worksheet, row, column, string=False):
    values = []
    while worksheet.cell(row=row, column=column).value:
        if string:
            values.append(str(worksheet.cell(row, column).value))
        else:
            values.append(worksheet.cell(row, column).value)
        row += 1
    return values


def get_data(filepath, **coeffs):
    #  открываем нужный лист в выбранной в форме книге excel
    workbook = openpyxl.load_workbook(filename=str(filepath))
    worksheet = workbook.worksheets[0]
    #  получаем массивы данных с листа по столбцам
    alpha = get_values(worksheet, 2, 2)
    beta = get_values(worksheet, 2, 3)
    v = get_values(worksheet, 2, 4)
    V = get_values(worksheet, 2, 5)
    teta = get_values(worksheet, 2, 6, string=True)
    k = [a / b for a, b in zip(V, v)]
    # убеждаемся, что массивы одинаковой длины
    assert len(teta) == len(v) == len(V) == len(alpha) == len(beta)
    #  приводим в нужный формат коэффициенты с формы
    T = int(coeffs['T'])
    F = float(coeffs['F'])
    sort_type = coeffs['sort']
    # интегрируем тета
    teta_integrated = [float(integrate(eval(teta[i]), (Symbol('x'), 0, T))) for i in range(0, len(teta))]
    #  вызываем функцию-решатель
    return [alpha, beta, v, teta_integrated, k, F], sort_type, workbook, worksheet


def sort_data(alpha, beta, v, teta_integrated, k, F, sort_type):
    if sort_type == 'β/α (max -> min)':
        ba = [b / a for a, b in zip(alpha, beta)]
        zipped = list(zip(ba, teta_integrated, alpha, beta, k, v))
        zipped.sort(reverse=True)
        ba, teta_integrated, alpha, beta, k, v = zip(*zipped)
    else:
        zipped = list(zip(teta_integrated, alpha, beta, k, v))
        zipped.sort(reverse=True)
        teta_integrated, alpha, beta, k, v = zip(*zipped)
    return [alpha, beta, v, teta_integrated, k, F]


def form_problem(alpha, beta, v, teta_integrated, k, F):
    n = len(alpha)
    problem = LpProblem('Zadachka', LpMaximize)
    x = LpVariable.dicts('x', range(n), lowBound=0, cat=LpContinuous)
    sum_var1 = lpSum([x[i] * v[i] * beta[i] for i in range(0, n)])
    sum_var2 = lpSum([x[i] * v[i] * alpha[i] for i in range(0, n)])
    problem += sum_var1 - sum_var2  # 'Функция цели "11.1"'
    problem += sum_var2 <= F  # "11.2"
    constraint1 = [x[i] <= k[i] for i in range(0, n)]
    for cnstr in constraint1:
        problem += cnstr
    constraint2 = [x[i] * v[i] <= teta_integrated[i] for i in range(0, n)]
    for cnstr in constraint2:
        problem += cnstr
    constraint3 = [teta_integrated[i] <= v[i] * (x[i] + 1) for i in range(0, n)]
    for cnstr in constraint3:
        problem += cnstr

    solved_prob = solve_problem(problem)

    return [solved_prob.variables(), pulp.LpStatus[solved_prob.status],
                    pulp.value(solved_prob.objective), solved_prob.solutionTime]


class Solved(object):
    def __init__(self, problem, number, variables, value, not_int_var, optimal=False):
        self.problem = problem
        self.number = number
        self.variables = variables
        self.value = value
        self.not_int_var = not_int_var
        self.optimal = optimal

    def __repr__(self):
        return self.problem

    def __str__(self):
        return self.problem

def solve_problem(problem):
    queue = {}
    acc = 0
    max_z = 0
    problem_copy = problem.deepcopy()
    problem_copy.solve()  # решаем задачу 0
    if problem_copy.status != 1:
        return problem  # возвращаем задачу, если она не оптимальна
    else:
        for v in problem_copy.variables():  # проверяем все х на целочисленность
            if v.varValue != int(v.varValue):
                queue += Solved(problem, acc, problem_copy.variables,
                                pulp.value(problem_copy.objective), v)  # добавляем первую проблему в очередь
                branch_and_bound(queue, max_z, acc)  # шаг 3
        return problem  # возвращаем проблему, если она целочислена


def branch_and_bound(queue, max_z, acc):  # передаем сюда задачу на ветвление, в том числе нецелую переменную
    if queue:
        max_z_queue = 0
        for prob in queue:
            if prob.value > max_z_queue:
                max_prob = prob
                max_z_queue = prob.value

        problem_left = max_prob.deepcopy()
        problem_left += max_prob.not_int_var <= floor(max_prob.not_int_var.varValue)
        problem_left_copy = problem_left.deepcopy()
        problem_left_copy.solve()
        acc += 1
        if problem_left_copy.status == 1:
            queue[acc] = problem_left

        problem_right = problem.deepcopy()
        problem_right += v >= ceil(v.varValue)
        problem_right_copy = problem_right.deepcopy()
        problem_right_copy.solve()
        acc += 1
        if problem_right_copy.status == 1:
            queue[acc] = problem_right

        if pulp.value(problem_right_copy.objective) > pulp.value(problem_left_copy.objective):
            for v in problem_right_copy.variables():  # проверяем все х на целочисленность
                if v.varValue != int(v.varValue):
                    branch_and_bound(problem_right, queue, max_z, acc, v, optimal)  # шаг 3
            max_z = pulp.value(problem_right_copy.objective)
            optimal[acc] = problem_right_copy
        else:
            for v in problem_right_copy.variables():  # проверяем все х на целочисленность
                if v.varValue != int(v.varValue):
                    branch_and_bound(problem_left, queue, max_z, acc - 1, v)  # шаг 3
            max_z = pulp.value(problem_left_copy.objective)
            if not optimal:
                optimal[acc] = problem_left_copy


    else:
        pass  # тут шаг 7


def show_results(variables, status, solution, time, sort_type):
    xs = []
    for v in variables:
        xs.append(str(v.name) + " = " + str(v.varValue))
    status = 'Статус: ' + status
    solution = 'Значение целевой функции: ' + str(solution)
    time = 'Время решения: ' + str(time) + ' сек.'
    sort_type = 'Сортировка: ' + sort_type
    results = [status, sort_type, solution, *xs, time]
    return results


def integer_lp(filepath, **coeffs):
    data, sort_type, workbook, worksheet = get_data(filepath, **coeffs)
    sorted_data = sort_data(*data, sort_type)
    problem = form_problem(*sorted_data)
    return show_results(*problem, sort_type)


def main():
    print(integer_lp('Zadachka (копия).xlsx', T=30, F=100000, sort='β/α (max -> min)'))


if __name__ == '__main__':
    main()

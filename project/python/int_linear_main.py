# -*- coding: utf-8 -*-
from anytree import Node, RenderTree
from anytree.exporter import DotExporter
from openpyxl import load_workbook
from pulp import *
from math import ceil, floor
from sympy import Symbol, integrate


class Solved(object):
    tree = []
    statuses = ['Не решено', 'Оптимально', 'Неопределенно', 'Не ограниченно', 'Нерешаемо']

    def nodenamefunc(node):
        node_params = [str(node.name), node.status, *node.xs]
        return "\n".join(node_params)

    def make_node(self):
        status = Solved.statuses[self.status]
        xs = [str(x[0]) + ' = ' + str(x[1]) for x in zip(self.problem.variables(), self.vars_value)]
        new_node = Node(name=self.number, status=status, xs=xs, parent_name=self.parent_number)
        for node in self.tree:
            if node.name == new_node.parent_name:
                new_node.parent = node
        self.tree.append(new_node)

    def __init__(self, problem, number, func_value, vars_value, status,
                 cont_var=None, cont_var_value=None, parent_number=None):
        self.problem = problem
        self.status = status
        self.number = number
        self.func_value = func_value
        self.vars_value = vars_value
        self.cont_var = cont_var
        self.cont_var_value = cont_var_value
        self.parent_number = parent_number
        self.make_node()

    def __repr__(self):
        return str(self.number)


class Solution(object):
    def __init__(self, status, acc, solution=None):
        self.status = status
        self.acc = acc
        self.has_sol = False
        if solution is not None:
            self.has_sol = True
            self.variables = solution.problem.variables()
            self.func_value = solution.func_value
            self.vars_value = solution.vars_value
            self.number = solution.number


def create_Solved(problem, acc, parent_number=None):
    problem_copy = problem.deepcopy()
    problem_copy.solve()
    acc += 1
    # создаем объект решенной задачи
    solved = Solved(problem=problem,
                    status=problem_copy.status,
                    number=acc,
                    func_value=value(problem_copy.objective),
                    vars_value=[var.varValue for var in problem_copy.variables()],
                    parent_number=parent_number)
    for v in problem_copy.variables():
        if v.varValue != int(v.varValue):
            solved.cont_var, solved.cont_var_value = v, v.varValue
    return solved, acc


def get_values(worksheet, row, column, expression=False):
    values = []
    while worksheet.cell(row=row, column=column).value:
        if expression:
            values.append(str(worksheet.cell(row, column).value))
        else:
            values.append(worksheet.cell(row, column).value)
        row += 1
    return values


def get_data_excel(filepath, T):
    #  открываем нужный лист в выбранной в форме книге excel
    workbook = load_workbook(filename=str(filepath))
    worksheet = workbook.worksheets[0]
    #  получаем массивы данных с листа по столбцам
    alpha = get_values(worksheet, 2, 2)
    beta = get_values(worksheet, 2, 3)
    v = get_values(worksheet, 2, 4)
    V = get_values(worksheet, 2, 5)
    teta = get_values(worksheet, 2, 6, expression=True)
    k = [a / b for a, b in zip(V, v)]
    # убеждаемся, что массивы одинаковой длины
    assert len(teta) == len(v) == len(V) == len(alpha) == len(beta)
    # интегрируем тета
    x = Symbol('x')
    teta_integrated = [float(integrate(eval(teta[i]), (x, 0, T))) for i in range(0, len(teta))]
    #  вызываем функцию-решатель
    return [alpha, beta, v, teta_integrated, V], workbook, worksheet


def sort_data(sort_type, alpha, beta, v, teta_integrated, V):
    if sort_type == 'b/a':
        ba = [b / a for a, b in zip(alpha, beta)]
        zipped = list(zip(ba, teta_integrated, alpha, beta, V, v))
        zipped.sort(reverse=True)
        ba, teta_integrated, alpha, beta, V, v = zip(*zipped)
    elif sort_type == 'teta':
        zipped = list(zip(teta_integrated, alpha, beta, V, v))
        zipped.sort(reverse=True)
        teta_integrated, alpha, beta, V, v = zip(*zipped)
    return [alpha, beta, v, teta_integrated, V]


def form_problem(alpha, beta, v, teta_integrated, V, **coeffs):

    def form_problem_1(alpha, beta, v, teta_integrated, V, F):
        k = [a / b for a, b in zip(V, v)]
        n = len(alpha)
        problem = LpProblem('Zadachka', LpMaximize)
        x = LpVariable.dicts('x', range(n), lowBound=0, cat=LpContinuous)
        sum_var1 = lpSum([x[i] * v[i] * beta[i] for i in range(0, n)])
        sum_var2 = lpSum([x[i] * v[i] * alpha[i] for i in range(0, n)])

        problem += sum_var1 - sum_var2  # 'Функция цели "11.1"'

        problem += sum_var2 <= F  # "11.2"
        constraint1 = [x[i] <= k[i] for i in range(0, n)]
        for cnstr in constraint1:
            problem += cnstr
        constraint2 = [x[i] * v[i] <= min(teta_integrated[i], V[i]) for i in range(0, n)]
        for cnstr in constraint2:
            problem += cnstr
        constraint3 = [min(teta_integrated[i], V[i]) <= v[i] * (x[i] + 1) for i in range(0, n)]
        for cnstr in constraint3:
            problem += cnstr
        problem.writeLP('results/ZLP_formula1')
        return problem

    def form_problem_2(alpha, beta, v, teta_integrated, V, F, D, y):
        k = [a / b for a, b in zip(V, v)]
        n = len(alpha)
        V = [v[i] * k[i] for i in range(0, n)]
        problem = LpProblem('Zadachka', LpMaximize)
        x = LpVariable.dicts('x', range(n), lowBound=0, cat=LpContinuous)
        sum_xvb = lpSum([x[i] * v[i] * beta[i] for i in range(0, n)])
        sum_xva = lpSum([x[i] * v[i] * alpha[i] for i in range(0, n)])
        sum_1yax = lpSum([(1 + y) * alpha[i] * x[i] for i in range(0, n)])
        sum_bx = lpSum([beta[i] * x[i] for i in range(0, n)])

        problem += sum_xvb + ((1 + y) * (F - sum_xva))  # цель

        problem += sum_xva <= F + D
        constraint1 = [x[i] <= k[i] for i in range(0, n)]
        constraint2 = [x[i] * v[i] <= min(teta_integrated[i], V[i]) for i in range(0, n)]
        constraint3 = [min(teta_integrated[i], V[i]) <= v[i] * (x[i] + 1) for i in range(0, n)]
        for cnstr in [*constraint1, *constraint2, *constraint3]:
            problem += cnstr
        problem += sum_1yax <= sum_bx

        problem.writeLP('results/ZLP_formula2')
        return problem

    if coeffs['zadacha'] == 1:
        return form_problem_1(alpha, beta, v, teta_integrated, V, coeffs['F'])
    elif coeffs['zadacha'] == 2:
        return form_problem_2(alpha, beta, v, teta_integrated, V, coeffs['F'], coeffs['D'], coeffs['y'])


def solve_problem(problem):
    # обнуляем дерево, потому что в одном прогоне программы может решаться несколько задач
    Solved.tree = []
    queue = []  # очередь на ветвление
    optimal = []  # лист оптимальных целочисленных решений
    acc = 0  # счетчик задач
    max_z = 0  # максимальное значение целевой функции

    first, acc = create_Solved(problem, acc, parent_number=None)
    # возвращаем задачу, если она не оптимальна
    if first.status != 1:
        return Solution(status='Нерешаемо', acc=acc, solution=None)
    else:
        # проверяем проблему на целочисленность
        if first.cont_var is not None:
            # добавляем первую проблему в очередь на ветвление
            queue.append(first)
            return branch_and_bound(queue, max_z, acc, optimal)
        # если проблема целочислена на первом шаге
        else:
            return Solution(acc=acc, status='Оптимально', solution=first)


def branch_and_bound(queue, max_z, acc, optimal):  # передаем сюда задачу на ветвление, в том числе нецелую переменную
    if queue:
        max_prob = queue[0]
        for prob in queue[1:]:
            if prob.value > max_prob.value:
                max_prob = prob
        queue.remove(max_prob)
        i = 1
        queue, max_z, acc, optimal = make_branch(max_prob, acc, queue, max_z, optimal, i)
        return branch_and_bound(queue, max_z, acc, optimal)
    else:
        if optimal:
            solution = optimal[0]
            for prob in optimal[1:]:
                if prob.func_value > solution.func_value:
                    solution = prob
            return Solution(acc=acc, status='Оптимально', solution=solution)
        else:
            return Solution(status='Нерешаемо', acc=acc, solution=None)


def make_branch(parent_problem, acc, queue, max_z, optimal, i):
    child_problem = parent_problem.problem.deepcopy()
    if i == 1:
        child_problem += parent_problem.cont_var <= floor(parent_problem.cont_var_value)  # левая ветвь
    elif i == 2:
        child_problem += parent_problem.cont_var >= ceil(parent_problem.cont_var_value)  # правая ветвь
    child_solved, acc = create_Solved(child_problem, acc, parent_number=parent_problem.number)
    if child_solved.status == 1:  # формируем подпроблему в очередь
        if child_solved.cont_var is None and max_z == 0:  # шаг 5
            max_z = child_solved.func_value
            optimal.append(child_solved)
            for prob in queue:  # проверяем, будем ли ветвить эту задачу
                if prob.value > max_z:
                    queue.append(child_solved)
                    break

        elif child_solved.cont_var is None and child_solved.func_value >= max_z:
            optimal.append(child_solved)

        elif child_solved.cont_var is not None and max_z == 0:
            queue.append(child_solved)

    if i == 1:
        return make_branch(parent_problem, acc, queue, max_z, optimal, i + 1)  # ветвим второй раз
    elif i == 2:
        return queue, max_z, acc, optimal


def write_to_excel(workbook, worksheet, filepath, sort_type, problem):
    for row in worksheet['H2:H9']:
        for cell in row:
            cell.value = None
    if not problem.has_sol:
        worksheet.cell(2, 8).value = 'Статус: ' + problem.status
        worksheet.cell(3, 8).value = 'Количество решенных ЗЛП: ' + str(problem.acc)
    else:
        worksheet.cell(2, 8).value = 'Статус: ' + problem.status
        worksheet.cell(3, 8).value = 'Сортировка: ' + sort_type
        worksheet.cell(4, 8).value = 'Значение целевой функции: ' + str(problem.func_value)
        worksheet.cell(5, 8).value = 'Количество решенных ЗЛП: ' + str(problem.acc)
        worksheet.cell(6, 8).value = 'Номер оптимальной задачи: ' + str(problem.number)
        for i in range(len(problem.variables)):
            worksheet.cell(6 + i, 8).value = (str(problem.variables[i]) +
                                              " = " + str(problem.vars_value[i]))
    workbook.save(filepath)


def show_results(sort_type, solved):
    if not solved.has_sol:
        status = 'Статус: ' + solved.status
        acc = 'Кол-во решенных ЗЛП: ' + str(solved.acc)
        results = [status, acc]
    else:
        status = 'Статус: ' + solved.status
        xs = [str(x[0]) + ' = ' + str(x[1]) for x in zip(solved.variables, solved.vars_value)]
        number_of_optimal = 'Номер оптимальной задачи: ' + str(solved.number)
        func_value = 'Значение целевой функции: ' + str(solved.func_value)
        sort_type = 'Сортировка: ' + sort_type
        acc = 'Кол-во решенных ЗЛП: ' + str(solved.acc)
        results = [status, func_value, *xs, sort_type, acc, number_of_optimal]
    DotExporter(Solved.tree[0], nodenamefunc=Solved.nodenamefunc).to_picture("results/ZLP_tree.png")
    return results


def integer_lp(filepath, **coeffs):
    data, workbook, worksheet = get_data_excel(filepath, coeffs['T'])
    sorted_data = sort_data(coeffs['sort'], *data)
    problem = form_problem(*sorted_data, **coeffs)
    solved_problem = solve_problem(problem)
    write_to_excel(workbook, worksheet, filepath, coeffs['sort'], solved_problem)
    return show_results(coeffs['sort'], solved_problem)


def main():
    print(integer_lp('Zadachka2.xlsx', T=1, F=30000, D=6000, y=0.125, zadacha=2, sort='b/a'))
    print(integer_lp('Zadachka.xlsx', T=30, F=100000, zadacha=1, sort='b/a'))


if __name__ == '__main__':
    main()

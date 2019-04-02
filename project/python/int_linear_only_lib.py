# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from pulp import *
from sympy import integrate, Symbol


def get_values(worksheet, row, column, expression=False):
    values = []
    while worksheet.cell(row=row, column=column).value:
        if expression:
            values.append(str(worksheet.cell(row, column).value))
        else:
            values.append(worksheet.cell(row, column).value)
        row += 1
    return values


def get_data_excel(filepath, T):
    #  открываем нужный лист в выбранной в форме книге excel
    workbook = load_workbook(filename=str(filepath))
    worksheet = workbook.worksheets[0]
    #  получаем массивы данных с листа по столбцам
    alpha = get_values(worksheet, 2, 2)
    beta = get_values(worksheet, 2, 3)
    v = get_values(worksheet, 2, 4)
    V = get_values(worksheet, 2, 5)
    teta = get_values(worksheet, 2, 6, expression=True)
    k = [a / b for a, b in zip(V, v)]
    # убеждаемся, что массивы одинаковой длины
    assert len(teta) == len(v) == len(V) == len(alpha) == len(beta)
    # интегрируем тета
    x = Symbol('x')
    teta_integrated = [float(integrate(eval(teta[i]), (x, 0, T))) for i in range(0, len(teta))]
    #  вызываем функцию-решатель
    return [alpha, beta, v, teta_integrated, k], workbook, worksheet


def sort_data(sort_type, alpha, beta, v, teta_integrated, k):
    if sort_type == 'b/a':
        ba = [b / a for a, b in zip(alpha, beta)]
        zipped = list(zip(ba, teta_integrated, alpha, beta, k, v))
        zipped.sort(reverse=True)
        ba, teta_integrated, alpha, beta, k, v = zip(*zipped)
    elif sort_type == 'teta':
        zipped = list(zip(teta_integrated, alpha, beta, k, v))
        zipped.sort(reverse=True)
        teta_integrated, alpha, beta, k, v = zip(*zipped)
    return [alpha, beta, v, teta_integrated, k]


def solve_problem(alpha, beta, v, teta_integrated, k, **coeffs):

    def solve_problem_1(alpha, beta, v, teta_integrated, k, F):
        n = len(alpha)
        problem = LpProblem('Zadachka', LpMaximize)
        x = LpVariable.dicts('x', range(n), lowBound=0, cat=LpInteger)
        sum_xvb = lpSum([x[i] * v[i] * beta[i] for i in range(0, n)])
        sum_xva = lpSum([x[i] * v[i] * alpha[i] for i in range(0, n)])
        problem += sum_xvb - sum_xva  # 'Функция цели "11.1"'
        problem += sum_xva <= F  # "11.2"
        constraint1 = [x[i] <= k[i] for i in range(0, n)]
        for cnstr in constraint1:
            problem += cnstr
        constraint2 = [x[i] * v[i] <= teta_integrated[i] for i in range(0, n)]
        for cnstr in constraint2:
            problem += cnstr
        constraint3 = [teta_integrated[i] <= v[i] * (x[i] + 1) for i in range(0, n)]
        for cnstr in constraint3:
            problem += cnstr

        problem.solve()
        problem.writeLP('formula1')
        return [problem.variables(), pulp.LpStatus[problem.status],
                pulp.value(problem.objective), problem.solutionTime]

    def solve_problem_2(alpha, beta, v, teta_integrated, k, F, D, y):
        n = len(alpha)
        V = [v[i] * k[i] for i in range(0, n)]
        problem = LpProblem('Zadachka', LpMaximize)
        x = LpVariable.dicts('x', range(n), lowBound=0, cat=LpInteger)
        sum_xvb = lpSum([x[i] * v[i] * beta[i] for i in range(0, n)])
        sum_xva = lpSum([x[i] * v[i] * alpha[i] for i in range(0, n)])
        sum_1yax = lpSum([(1 + y) * alpha[i] * x[i] for i in range(0, n)])
        sum_bx = lpSum([beta[i] * x[i] for i in range(0, n)])

        problem += sum_xvb + ((1 + y) * (F - sum_xva))  # цель

        problem += sum_xva <= F + D
        constraint1 = [x[i] <= k[i] for i in range(0, n)]
        constraint2 = [x[i] * v[i] <= min(teta_integrated[i], V[i]) for i in range(0, n)]
        constraint3 = [min(teta_integrated[i], V[i]) <= v[i] * (x[i] + 1) for i in range(0, n)]
        for cnstr in [*constraint1, *constraint2, *constraint3]:
            problem += cnstr
        problem += sum_1yax <= sum_bx

        problem.solve()
        problem.writeLP('formula2')
        return [problem.variables(), pulp.LpStatus[problem.status],
                pulp.value(problem.objective), problem.solutionTime]

    if coeffs['zadacha'] == 1:
        return solve_problem_1(alpha, beta, v, teta_integrated, k,  coeffs['F'])
    elif coeffs['zadacha'] == 2:
        return solve_problem_2(alpha, beta, v, teta_integrated, k, coeffs['F'], coeffs['D'], coeffs['y'])


def show_results(sort_type, variables, status, solution, time):
    status = 'Статус: ' + status
    solution = 'Значение целевой функции: ' + str(solution)
    time = 'Время решения: ' + str(time) + ' сек.'
    sort_type = 'Сортировка: ' + sort_type
    results = [status, sort_type, solution, time]
    for v in variables:
        results.append(str(v.name) + " = " + str(v.varValue))
    return results


def write_to_excel(workbook, worksheet, filepath, sort_type, *problem):
    variables, status, solution, time = problem
    worksheet.cell(2, 8).value = 'Статус: ' + status
    worksheet.cell(3, 8).value = 'Сортировка: ' + sort_type
    worksheet.cell(4, 8).value = 'Значение целевой функции: ' + str(solution)
    for i in range(len(variables)):
        worksheet.cell(5 + i, 8).value = (str(variables[i].name) +
                                          " = " + str(variables[i].varValue))
    worksheet.cell(6 + i, 8).value = 'Время решения: ' + str(time) + ' сек.'
    workbook.save(filepath)


def integer_lp(filepath, **coeffs):
    data, workbook, worksheet = get_data_excel(filepath, coeffs['T'])
    sorted_data = sort_data(coeffs['sort'], *data)
    problem = solve_problem(*sorted_data, **coeffs)
    #  write_to_excel(workbook, worksheet, filepath, coeffs['sort'], *problem)
    return show_results(coeffs['sort'], *problem)


def main():
    print(integer_lp('Zadachka2.xlsx', T=1, F=30000, D=12000, y=0.11, zadacha=2, sort='β/α (max -> min)'))
    print(integer_lp('Zadachka.xlsx', T=30, F=100000, zadacha=1, sort='β/α (max -> min)'))


if __name__ == '__main__':
    main()
